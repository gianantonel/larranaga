"""
R-10: Generación del archivo HWCRARCA para importar a Holistor.

Toma el DataFrame procesado (post R-01 + R-02) y genera un .xlsx con la
estructura que Holistor espera, listo para que el usuario lo abra y lo guarde
como .prn (texto delimitado por espacios) para importarlo en su módulo IVA.

Estructura del archivo de salida (pestaña "HWComprobantes Recibidos"):
  Fila 1: Banner "Datos origen copia desde ARCA - MIS COMPROBANTES RECIBIDOS"
  Fila 2: Headers (30 columnas A-AD según convención Holistor)
  Fila 3+: Datos transformados

Transformaciones aplicadas respecto del output de R-02:
  - Tipo: "1 - Factura A" → "1" (solo código numérico)
  - Tipo Cambio: "1" → 1.00 (numérico con 2 decimales)
  - Tipo Doc: "CUIT" → "80", "DNI" → "96", vacío → "99"
  - Consumidor Final: si CUIT está vacío/0 → "0" + "CONSUMIDOR FINAL" + tipo doc 99
  - Razón social vacía: "SIN DENOMINACION"
  - Validación: chequea que Imp.Total ≈ Σ(componentes) ± 0.01

Referencias:
  - "Instructivo Holistor — Importación de Mis Comprobantes Recibidos ARCA"
  - Fixture HWCRARCA.xls (template oficial de Holistor)
"""

import io
import re
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .division_alicuotas import parse_string_float


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────

# Headers de las 48 columnas A-AV según el template oficial de Holistor
# Cols 0-29 (A-AD): datos directos de ARCA
# Cols 30-34 (AE-AI): totales recalculados (validación interna)
# Cols 35-36 (AJ-AK): metadata adicional (Fecha Recepción, Pcia)
# Cols 37-45 (AL-AT): imputación contable (requiere Maestro de Proveedores — R-09)
# Col  46 (AU): Moneda secundaria
# Col  47 (AV): DIFERENCIA = Importe Total ARCA − Importe Total recalculado
HWCRARCA_HEADERS: List[str] = [
    "Fecha Emisión",                           # A
    "Cpbte",                                   # B
    "Suc.",                                    # C
    "Número",                                  # D
    "N° Hasta",                                # E
    "Cód. Autorización",                       # F
    "Tipo Doc.",                               # G
    "CUIT",                                    # H
    "Razón Social/Denominación Proveedor",     # I
    "Tipo Doc. Receptor",                      # J
    "Nro. Doc. Receptor",                      # K
    "Tipo Cbio.",                              # L
    "Moneda",                                  # M
    "Neto Gravado",                            # N
    "IVA 2,5%",                                # O
    "Imp. Neto Gravado IVA 2,5%",              # P
    "IVA 5%",                                  # Q
    "Imp. Neto Gravado IVA 5%",                # R
    "IVA 10,5%",                               # S
    "Imp. Neto Gravado IVA 10,5%",             # T
    "IVA 21%",                                 # U
    "Imp. Neto Gravado IVA 21%",               # V
    "IVA 27%",                                 # W
    "Imp. Neto Gravado IVA 27%",               # X
    "Imp. Neto Gravado Total",                 # Y
    "Importes No Gravados",                    # Z
    "Importes Exentos",                        # AA
    "Otros Tributos",                          # AB
    "Total IVA",                               # AC
    "Importe   Total",                         # AD — Holistor usa doble espacio
    # ── Totales recalculados (validación interna Holistor) ─────────────────
    "Neto Gravado",                            # AE — duplicado de Y para chequeo
    "Importes No Gravados",                    # AF — duplicado de Z
    "Importes Exentos",                        # AG — duplicado de AA
    "IVA  Liquidado",                          # AH — Holistor usa doble espacio
    "Importe   Total",                         # AI — Holistor usa doble espacio
    # ── Metadata adicional ─────────────────────────────────────────────────
    "Fecha Recepción",                         # AJ — default = Fecha Emisión
    "Pcia",                                    # AK — vacío (sin maestro)
    # ── Imputación contable (R-09 — requiere Maestro de Proveedores) ──────
    "Neto Cód.",                               # AL
    "Alíc.",                                   # AM
    "Neto a importar",                         # AN
    "IVA   Crédito",                           # AO — Holistor usa triple espacio
    "NG/EX  Cód.",                             # AP — Holistor usa doble espacio
    "P/R Cód.",                                # AQ
    "Perc./Ret.",                              # AR
    "Pcia P/R",                                # AS
    "Cód Cte",                                 # AT
    # ── Moneda secundaria + DIFERENCIA de cuadre ──────────────────────────
    "Moneda",                                  # AU
    "DIFERENCIA",                              # AV — Importe Total ARCA − recalculado
]

# Mapeo R-02 → posición HWCRARCA (índices A-AD = 0-29)
ARCA_COLS_ORDERED: List[str] = [
    "Fecha", "Tipo", "Punto de Venta", "Número Desde", "Número Hasta",
    "Cód. Autorización", "Tipo Doc. Emisor", "Nro. Doc. Emisor",
    "Denominación Emisor", "Tipo Doc. Receptor", "Nro. Doc. Receptor",
    "Tipo Cambio", "Moneda",
    # Columnas N-AD (numéricas)
    "Neto Grav. IVA 0%",
    "IVA 2,5%", "Neto Grav. IVA 2,5%",
    "IVA 5%", "Neto Grav. IVA 5%",
    "IVA 10,5%", "Neto Grav. IVA 10,5%",
    "IVA 21%", "Neto Grav. IVA 21%",
    "IVA 27%", "Neto Grav. IVA 27%",
    "Neto Gravado Total", "Neto No Gravado", "Op. Exentas",
    "Otros Tributos", "Total IVA", "Imp. Total",
]

# Índices de columnas que deben emitirse como números (no strings)
NUMERIC_COL_INDICES = list(range(13, 30))  # N a AD: Neto, IVAs, Netos, Totales

# Mapeo de Tipo Doc texto → código numérico AFIP
TIPO_DOC_MAP: Dict[str, str] = {
    "CUIT": "80",
    "CUIL": "86",
    "DNI":  "96",
    "":     "99",
    "0":    "99",
    "NAN":  "99",
}

# Códigos de comprobantes Tipo B y C (no llevan IVA discriminado)
# Holistor: para estos tipos, Neto Gravado e IVA Liquidado deben ser 0
TIPOS_B_C = {
    "6", "7", "8", "9",                          # B básicos
    "11", "12", "13", "15",                      # C básicos
    "206", "207", "208",                          # B MiPyME
    "211", "212", "213",                          # C MiPyME
}

# Banner de la fila 1 del template Holistor
BANNER_TEXT = "Datos origen copia desde ARCA - MIS COMPROBANTES RECIBIDOS"


# ─────────────────────────────────────────────────────────────────────────────
# TRANSFORMACIONES
# ─────────────────────────────────────────────────────────────────────────────

def extraer_codigo_tipo(tipo_str: Any) -> str:
    """'1 - Factura A' → '1' · '11 - Factura C' → '11' · vacío → ''"""
    if pd.isna(tipo_str):
        return ""
    s = str(tipo_str).strip()
    if not s:
        return ""
    m = re.match(r"^\s*(\d+)\s*-", s)
    if m:
        return m.group(1)
    # Si ya es solo un número, devolverlo
    if s.isdigit():
        return s
    return ""


def formatear_tipo_cambio(valor: Any) -> float:
    """Devuelve siempre un float. '1' → 1.0 · '1,00' → 1.0 · vacío → 1.0"""
    if pd.isna(valor) or str(valor).strip() == "":
        return 1.0
    s = str(valor).strip()
    if s in {"$", ""}:
        return 1.0
    try:
        return parse_string_float(s) or 1.0
    except Exception:
        return 1.0


def mapear_tipo_doc(valor: Any) -> str:
    """'CUIT' → '80' · 'DNI' → '96' · vacío → '99' · ya numérico → tal cual"""
    if pd.isna(valor):
        return "99"
    s = str(valor).strip().upper()
    # Si ya es numérico (e.g. "80", "96"), devolverlo
    if s.isdigit():
        return s
    return TIPO_DOC_MAP.get(s, "99")


def normalizar_consumidor_final(row: pd.Series) -> pd.Series:
    """
    Si CUIT vacío/0 → CUIT="0", Razón Social="CONSUMIDOR FINAL", Tipo Doc=99.
    Si solo Razón Social vacía → "SIN DENOMINACION".
    """
    cuit  = str(row.get("Nro. Doc. Emisor", "")).strip()
    razon = str(row.get("Denominación Emisor", "")).strip()

    cuit_vacio = cuit in {"", "0", "nan", "NaN", "None"}
    if cuit_vacio:
        row["Nro. Doc. Emisor"]   = "0"
        row["Denominación Emisor"] = "CONSUMIDOR FINAL"
        row["Tipo Doc. Emisor"]   = "99"
    elif razon in {"", "nan", "NaN", "None"}:
        row["Denominación Emisor"] = "SIN DENOMINACION"

    return row


def aplicar_reglas_tipo_b_c(row: pd.Series) -> pd.Series:
    """
    Para comprobantes Tipo B/C, Holistor exige Neto Gravado e IVA = 0.
    Solo el Importe Total debe tener valor.
    R-01 ya hace gran parte; acá lo reforzamos para asegurar el formato Holistor.
    """
    tipo_codigo = extraer_codigo_tipo(row.get("Tipo", ""))
    if tipo_codigo not in TIPOS_B_C:
        return row

    # Forzar a 0 las columnas de neto e IVA por alícuota
    cols_a_cero = [
        "Neto Grav. IVA 0%",
        "IVA 2,5%", "Neto Grav. IVA 2,5%",
        "IVA 5%", "Neto Grav. IVA 5%",
        "IVA 10,5%", "Neto Grav. IVA 10,5%",
        "IVA 21%", "Neto Grav. IVA 21%",
        "IVA 27%", "Neto Grav. IVA 27%",
        "Neto Gravado Total", "Total IVA",
    ]
    for col in cols_a_cero:
        if col in row.index:
            row[col] = "0"
    return row


# ─────────────────────────────────────────────────────────────────────────────
# VALIDACIÓN DE CUADRE
# ─────────────────────────────────────────────────────────────────────────────

class CuadreError(Exception):
    """Excepción que se lanza cuando los totales agregados Debe ≠ Haber.

    El endpoint captura esta excepción y la traduce a HTTP 422 con el
    detalle estructurado en `result` para que el frontend muestre dónde
    rompe el cuadre.
    """
    def __init__(self, mensaje: str, result: "Dict[str, Any]"):
        super().__init__(mensaje)
        self.result = result


def validar_cuadre(
    df: pd.DataFrame,
    tolerancia: float = 0.01,
    tolerancia_pct: float = 1.0,
) -> Dict[str, Any]:
    """
    Hook de validación pre-exportación HWCRARCA.

    Verifica DOS niveles:

    1. AGREGADO (Debe = Haber a nivel contable, BLOQUEANTE):
         Σ Imp.Total ≈ Σ (Neto Grav + No Gravados + Exentas + Total IVA + Otros Tributos
                          + Imp.Total de Tipo B/C, que Holistor trata como No Gravado)
       Si la diferencia agregada > max(tolerancia, haber * tolerancia_pct/100),
       retorna `cuadre_ok=False` y el archivo NO debe exportarse.

    2. POR FILA (informativo, NO BLOQUEANTE):
         Para cada fila chequea coherencia interna. Las diferencias se reportan
         como `advertencias` pero no detienen la exportación.

    Tolerancia dual:
      - Absoluta (default $0.01): cubre errores de redondeo en archivos chicos.
      - Relativa (default 1.0%): cubre archivos grandes donde ARCA no expone
        percepciones discriminadas (caso típico BUTALO Feb 2026: dif 0.05%).
      Se acepta si la diferencia es <= MAX(tolerancia, haber * tolerancia_pct/100).

    Args:
        df: DataFrame post R-01/R-02
        tolerancia: diferencia absoluta aceptable (default $0.01)
        tolerancia_pct: diferencia relativa aceptable en % del haber (default 1.0)

    Returns: dict con
        - cuadre_ok: bool (basado en agregados)
        - debe_total / haber_total / diferencia_agregada / diferencia_pct
        - tolerancia / tolerancia_pct / tolerancia_efectiva
        - totales: dict con desglose
        - advertencias / errores
        - total_filas / filas_con_advertencia
        - valido: alias retrocompatible de cuadre_ok
    """
    errores: List[str] = []
    advertencias: List[str] = []

    suma_neto = 0.0
    suma_no_gravado = 0.0
    suma_exentas = 0.0
    suma_iva = 0.0
    suma_otros = 0.0
    suma_total = 0.0
    # Holistor trata el Imp.Total de Tipo B/C como No Gravado (instructivo: "Holistor
    # toma Otros tributos como No gravados"). Para que el cuadre agregado funcione,
    # acumulamos esos montos por separado para sumarlos al Debe contable.
    suma_b_c_imp_total = 0.0

    for idx, row in df.iterrows():
        try:
            tipo = extraer_codigo_tipo(row.get("Tipo", ""))
            es_b_c = tipo in TIPOS_B_C

            neto    = parse_string_float(row.get("Neto Gravado Total", "0"))
            no_grav = parse_string_float(row.get("Neto No Gravado",    "0"))
            exentas = parse_string_float(row.get("Op. Exentas",        "0"))
            iva     = parse_string_float(row.get("Total IVA",          "0"))
            otros   = parse_string_float(row.get("Otros Tributos",     "0"))
            total   = parse_string_float(row.get("Imp. Total",         "0"))

            # Agregar a sumas
            suma_neto       += neto
            suma_no_gravado += no_grav
            suma_exentas    += exentas
            suma_iva        += iva
            suma_otros      += otros
            suma_total      += total

            # Validación por fila (informativa)
            calculado = neto + no_grav + exentas + iva + otros
            diff = abs(calculado - total)

            # Tipo B/C: si solo Imp.Total tiene valor, OK individual
            # y el Imp.Total se imputa como No Gravado a nivel agregado
            if es_b_c and total > 0 and calculado < 0.01:
                suma_b_c_imp_total += total
                continue

            if diff > tolerancia:
                num = row.get("Número Desde", "?")
                advertencias.append(
                    f"Fila {idx + 1} (cbte {num}): "
                    f"Imp.Total={total:.2f} vs Σcomponentes={calculado:.2f} "
                    f"(dif={diff:.2f})"
                )
        except Exception as exc:
            errores.append(f"Fila {idx + 1}: error al validar — {exc}")

    # ── Validación AGREGADA — lo que determina si exportar o no ──
    # Debe = Σ componentes + Imp.Total de B/C (Holistor lo trata como No Gravado)
    debe_total      = (
        suma_neto + suma_no_gravado + suma_exentas + suma_iva + suma_otros
        + suma_b_c_imp_total
    )
    haber_total     = suma_total
    diff_agregada   = abs(debe_total - haber_total)
    diff_pct        = (diff_agregada / haber_total * 100.0) if haber_total > 0 else 0.0

    # Tolerancia dual: pasa si está dentro de la absoluta O de la relativa
    tol_relativa_abs = haber_total * tolerancia_pct / 100.0
    tol_efectiva     = max(tolerancia, tol_relativa_abs)
    cuadre_ok        = diff_agregada <= tol_efectiva and len(errores) == 0

    return {
        "cuadre_ok":             cuadre_ok,
        "valido":                cuadre_ok,
        "debe_total":            round(debe_total, 2),
        "haber_total":           round(haber_total, 2),
        "diferencia_agregada":   round(diff_agregada, 2),
        "diferencia_pct":        round(diff_pct, 4),
        "tolerancia":            tolerancia,
        "tolerancia_pct":        tolerancia_pct,
        "tolerancia_efectiva":   round(tol_efectiva, 2),
        "totales": {
            "neto_gravado":    round(suma_neto, 2),
            "no_gravado":      round(suma_no_gravado, 2),
            "exentas":         round(suma_exentas, 2),
            "iva":             round(suma_iva, 2),
            "otros_tributos":  round(suma_otros, 2),
            "imp_total":       round(suma_total, 2),
            "b_c_no_grav_implicito": round(suma_b_c_imp_total, 2),
        },
        "errores":               errores,
        "advertencias":          advertencias,
        "total_filas":           len(df),
        "filas_con_advertencia": len(advertencias),
    }


# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUCCIÓN DEL ARCHIVO HWCRARCA
# ─────────────────────────────────────────────────────────────────────────────

def construir_hwcrarca_xlsx(df: pd.DataFrame) -> Tuple[bytes, Dict[str, Any]]:
    """
    Genera un .xlsx con la estructura HWCRARCA lista para importar a Holistor.

    El usuario debe:
      1. Abrir el archivo descargado.
      2. Guardar como .prn (Texto delimitado por espacios) — pestaña activa
         "HWComprobantes Recibidos".
      3. Importar en Holistor: Útiles → Importar Datos de otros sistemas →
         Compras → IVA Compras → seleccionar el .prn.

    Args:
        df: DataFrame del output de R-02 (post limpieza + división de alícuotas).

    Returns:
        (xlsx_bytes, stats):
          - xlsx_bytes: bytes del archivo .xlsx generado
          - stats: { total_filas, filas_consumidor_final, filas_b_c,
                     validacion: { valido, errores, advertencias, ... } }
    """
    df = df.copy()

    # 1. Transformaciones por celda
    if "Tipo" in df.columns:
        df["Tipo"] = df["Tipo"].apply(extraer_codigo_tipo)

    if "Tipo Doc. Emisor" in df.columns:
        df["Tipo Doc. Emisor"] = df["Tipo Doc. Emisor"].apply(mapear_tipo_doc)
    if "Tipo Doc. Receptor" in df.columns:
        df["Tipo Doc. Receptor"] = df["Tipo Doc. Receptor"].apply(mapear_tipo_doc)

    # 2. Reglas por fila
    df = df.apply(normalizar_consumidor_final, axis=1)
    df = df.apply(aplicar_reglas_tipo_b_c,    axis=1)

    # 3. Stats agregados (informativos)
    filas_consumidor_final = int((df["Denominación Emisor"] == "CONSUMIDOR FINAL").sum())
    filas_b_c              = int(df["Tipo"].isin(TIPOS_B_C).sum())

    # 4. Validación de cuadre — HOOK PRE-EXPORTACIÓN
    # Si los totales agregados NO cuadran (Debe ≠ Haber a nivel contable),
    # bloqueamos la exportación lanzando CuadreError. El endpoint la traduce a HTTP 422.
    validacion = validar_cuadre(df)
    if not validacion["cuadre_ok"]:
        raise CuadreError(
            f"El archivo NO cuadra contablemente: "
            f"Debe={validacion['debe_total']:.2f} vs Haber={validacion['haber_total']:.2f} "
            f"(diferencia={validacion['diferencia_agregada']:.2f}, "
            f"tolerancia={validacion['tolerancia']:.2f}). "
            f"Revise los datos antes de importar a Holistor.",
            result=validacion,
        )

    # 5. Construcción del Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "HWComprobantes Recibidos"

    # Fila 1: banner
    cell = ws.cell(row=1, column=1, value=BANNER_TEXT)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="305496")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HWCRARCA_HEADERS))
    ws.row_dimensions[1].height = 22

    # Fila 2: headers
    for i, header in enumerate(HWCRARCA_HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=header)
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Fila 3+: datos
    for row_pos, (_, row) in enumerate(df.iterrows(), start=3):
        # ── 1. Cols A-AD (0-29): datos directos de ARCA ─────────────────────
        for col_pos, arca_col in enumerate(ARCA_COLS_ORDERED):
            valor_raw = row.get(arca_col, "")

            # Tipo Cambio (col L = índice 11): siempre float
            if col_pos == 11:
                valor = formatear_tipo_cambio(valor_raw)
                c = ws.cell(row=row_pos, column=col_pos + 1, value=valor)
                c.number_format = "0.00"
                continue

            # Columnas numéricas (N-AD): convertir string ARCA a float
            if col_pos in NUMERIC_COL_INDICES:
                valor = parse_string_float(valor_raw)
                c = ws.cell(row=row_pos, column=col_pos + 1, value=valor)
                c.number_format = "#,##0.00"
                continue

            # Columnas texto: emitir como string, NaN → vacío
            if pd.isna(valor_raw):
                valor = ""
            else:
                valor = str(valor_raw).strip()
            ws.cell(row=row_pos, column=col_pos + 1, value=valor)

        # ── 2. Cols AE-AI (30-34): totales recalculados (duplicado validación) ──
        neto_gravado = parse_string_float(row.get("Neto Gravado Total", "0"))
        no_gravado   = parse_string_float(row.get("Neto No Gravado",    "0"))
        exentas      = parse_string_float(row.get("Op. Exentas",        "0"))
        total_iva    = parse_string_float(row.get("Total IVA",          "0"))
        imp_total    = parse_string_float(row.get("Imp. Total",         "0"))
        otros_trib   = parse_string_float(row.get("Otros Tributos",     "0"))

        for col_idx, val in [
            (31, neto_gravado),  # AE = col 31
            (32, no_gravado),    # AF
            (33, exentas),       # AG
            (34, total_iva),     # AH — IVA Liquidado
            (35, imp_total),     # AI
        ]:
            c = ws.cell(row=row_pos, column=col_idx, value=val)
            c.number_format = "#,##0.00"

        # ── 3. Cols AJ-AK (35-36): metadata ────────────────────────────────
        fecha_emision = row.get("Fecha", "")
        if not pd.isna(fecha_emision):
            ws.cell(row=row_pos, column=36, value=str(fecha_emision).strip())  # AJ Fecha Recepción
        ws.cell(row=row_pos, column=37, value="")  # AK Pcia — vacío sin maestro

        # ── 4. Cols AL-AT (38-46): imputación contable — VACÍAS (R-09 pendiente) ─
        # Cuando se cargue el Maestro de Proveedores, acá va:
        #   Neto Cód., Alíc., Neto a importar, IVA Crédito, NG/EX Cód.,
        #   P/R Cód., Perc./Ret., Pcia P/R, Cód Cte (cuenta contable por CUIT)
        for col_idx in range(38, 47):  # AL..AT (1-indexed: 38..46)
            ws.cell(row=row_pos, column=col_idx, value="")

        # ── 5. Col AU (46): Moneda secundaria (igual a la primera) ─────────
        moneda = row.get("Moneda", "PES")
        ws.cell(row=row_pos, column=47, value=str(moneda).strip() if not pd.isna(moneda) else "PES")

        # ── 6. Col AV (47): DIFERENCIA = ARCA Total − recalculado ──────────
        # Recalculado = Neto + NoGrav + Exentas + IVA + Otros
        recalculado = neto_gravado + no_gravado + exentas + total_iva + otros_trib
        diferencia  = round(imp_total - recalculado, 2)
        c = ws.cell(row=row_pos, column=48, value=diferencia)
        c.number_format = "#,##0.00"

    # Anchos de columna
    widths = {
        "A": 12, "B": 7,  "C": 7,  "D": 11, "E": 11, "F": 18,
        "G": 9,  "H": 13, "I": 32, "J": 9,  "K": 13, "L": 9,
        "M": 8,
    }
    for letra, ancho in widths.items():
        ws.column_dimensions[letra].width = ancho
    # Numéricas: ancho fijo
    for col_letter in [chr(ord("N") + i) if i < 13 else "A" + chr(ord("A") + i - 13)
                       for i in range(17)]:
        ws.column_dimensions[col_letter].width = 14

    # Freeze panes en fila 3 para scroll cómodo
    ws.freeze_panes = "A3"

    out = io.BytesIO()
    wb.save(out)

    stats = {
        "total_filas":            len(df),
        "filas_consumidor_final": filas_consumidor_final,
        "filas_b_c":              filas_b_c,
        "validacion":             validacion,
    }

    print(f"  -> HWCRARCA generado:")
    print(f"     - Filas:                {stats['total_filas']}")
    print(f"     - Tipo B/C:             {filas_b_c}")
    print(f"     - Consumidor final:     {filas_consumidor_final}")
    print(f"     - Cuadre OK:            {validacion['valido']}")
    print(f"     - Filas con advertencia: {validacion['filas_con_advertencia']}")

    return out.getvalue(), stats
