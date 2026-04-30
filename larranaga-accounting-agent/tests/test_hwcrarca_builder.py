"""Tests para R-10: hwcrarca_builder."""

import io
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.transformaciones.hwcrarca_builder import (  # noqa: E402
    BANNER_TEXT,
    HWCRARCA_HEADERS,
    aplicar_reglas_tipo_b_c,
    construir_hwcrarca_xlsx,
    extraer_codigo_tipo,
    formatear_tipo_cambio,
    mapear_tipo_doc,
    normalizar_consumidor_final,
    validar_cuadre,
)


# ─── extraer_codigo_tipo ──────────────────────────────────────────────────────

class TestExtraerCodigoTipo:
    def test_factura_a(self):
        assert extraer_codigo_tipo("1 - Factura A") == "1"

    def test_factura_c(self):
        assert extraer_codigo_tipo("11 - Factura C") == "11"

    def test_mipyme(self):
        assert extraer_codigo_tipo("201 - F.Crédito MiPyME A") == "201"

    def test_solo_numero(self):
        assert extraer_codigo_tipo("6") == "6"

    def test_vacio(self):
        assert extraer_codigo_tipo("") == ""
        assert extraer_codigo_tipo(None) == ""

    def test_nan(self):
        import math
        assert extraer_codigo_tipo(math.nan) == ""

    def test_sin_guion(self):
        assert extraer_codigo_tipo("Factura A") == ""


# ─── formatear_tipo_cambio ────────────────────────────────────────────────────

class TestFormatearTipoCambio:
    def test_entero(self):
        assert formatear_tipo_cambio("1") == 1.0

    def test_decimal_coma(self):
        # ARCA usa coma como separador decimal
        assert formatear_tipo_cambio("1,50") == 1.5

    def test_vacio(self):
        assert formatear_tipo_cambio("") == 1.0
        assert formatear_tipo_cambio(None) == 1.0

    def test_signo_dolar(self):
        # Dato real de ARCA con moneda en columna anterior
        assert formatear_tipo_cambio("$") == 1.0


# ─── mapear_tipo_doc ──────────────────────────────────────────────────────────

class TestMapearTipoDoc:
    def test_cuit(self):
        assert mapear_tipo_doc("CUIT") == "80"

    def test_cuil(self):
        assert mapear_tipo_doc("CUIL") == "86"

    def test_dni(self):
        assert mapear_tipo_doc("DNI") == "96"

    def test_vacio(self):
        assert mapear_tipo_doc("") == "99"
        assert mapear_tipo_doc(None) == "99"

    def test_ya_numerico(self):
        assert mapear_tipo_doc("80") == "80"
        assert mapear_tipo_doc("99") == "99"

    def test_desconocido(self):
        assert mapear_tipo_doc("XYZ") == "99"


# ─── normalizar_consumidor_final ──────────────────────────────────────────────

class TestNormalizarConsumidorFinal:
    def test_cuit_vacio(self):
        row = pd.Series({
            "Nro. Doc. Emisor":   "",
            "Denominación Emisor": "",
            "Tipo Doc. Emisor":   "CUIT",
        })
        out = normalizar_consumidor_final(row)
        assert out["Nro. Doc. Emisor"]   == "0"
        assert out["Denominación Emisor"] == "CONSUMIDOR FINAL"
        assert out["Tipo Doc. Emisor"]   == "99"

    def test_cuit_cero(self):
        row = pd.Series({
            "Nro. Doc. Emisor":   "0",
            "Denominación Emisor": "",
            "Tipo Doc. Emisor":   "CUIT",
        })
        out = normalizar_consumidor_final(row)
        assert out["Denominación Emisor"] == "CONSUMIDOR FINAL"
        assert out["Tipo Doc. Emisor"]   == "99"

    def test_cuit_valido_razon_vacia(self):
        row = pd.Series({
            "Nro. Doc. Emisor":   "30709212083",
            "Denominación Emisor": "",
            "Tipo Doc. Emisor":   "CUIT",
        })
        out = normalizar_consumidor_final(row)
        assert out["Nro. Doc. Emisor"]   == "30709212083"
        assert out["Denominación Emisor"] == "SIN DENOMINACION"

    def test_todo_completo_no_modifica(self):
        row = pd.Series({
            "Nro. Doc. Emisor":   "30709212083",
            "Denominación Emisor": "ACME S.A.",
            "Tipo Doc. Emisor":   "CUIT",
        })
        out = normalizar_consumidor_final(row.copy())
        assert out["Denominación Emisor"] == "ACME S.A."


# ─── aplicar_reglas_tipo_b_c ──────────────────────────────────────────────────

class TestReglasTipoBC:
    def test_factura_b_neto_iva_a_cero(self):
        row = pd.Series({
            "Tipo":                "6",
            "Neto Grav. IVA 21%":  "10000",
            "IVA 21%":             "2100",
            "Neto Gravado Total":  "10000",
            "Total IVA":           "2100",
            "Imp. Total":          "12100",
        })
        out = aplicar_reglas_tipo_b_c(row.copy())
        assert out["Neto Grav. IVA 21%"]  == "0"
        assert out["IVA 21%"]             == "0"
        assert out["Neto Gravado Total"]  == "0"
        assert out["Total IVA"]           == "0"
        # Imp. Total NO se toca
        assert out["Imp. Total"]          == "12100"

    def test_factura_a_no_modifica(self):
        row = pd.Series({
            "Tipo":                "1",
            "Neto Grav. IVA 21%":  "10000",
            "IVA 21%":             "2100",
        })
        out = aplicar_reglas_tipo_b_c(row.copy())
        assert out["Neto Grav. IVA 21%"]  == "10000"
        assert out["IVA 21%"]             == "2100"


# ─── validar_cuadre ───────────────────────────────────────────────────────────

class TestValidarCuadre:
    def test_factura_a_cuadra(self):
        df = pd.DataFrame([{
            "Tipo": "1 - Factura A",
            "Número Desde":      "991",
            "Neto Gravado Total": "10000",
            "Neto No Gravado":    "0",
            "Op. Exentas":        "0",
            "Total IVA":          "2100",
            "Otros Tributos":     "0",
            "Imp. Total":         "12100",
        }])
        v = validar_cuadre(df)
        assert v["valido"]
        assert v["filas_con_advertencia"] == 0

    def test_factura_b_imp_total_solo(self):
        # Tipo B: solo Imp. Total tiene valor, todo el resto en 0
        df = pd.DataFrame([{
            "Tipo":               "6 - Factura B",
            "Número Desde":      "100",
            "Neto Gravado Total": "0",
            "Neto No Gravado":    "0",
            "Op. Exentas":        "0",
            "Total IVA":          "0",
            "Otros Tributos":     "0",
            "Imp. Total":         "12100",
        }])
        v = validar_cuadre(df)
        assert v["valido"]
        assert v["filas_con_advertencia"] == 0

    def test_factura_a_no_cuadra(self):
        df = pd.DataFrame([{
            "Tipo":               "1 - Factura A",
            "Número Desde":      "991",
            "Neto Gravado Total": "10000",
            "Neto No Gravado":    "0",
            "Op. Exentas":        "0",
            "Total IVA":          "2100",
            "Otros Tributos":     "0",
            "Imp. Total":         "15000",  # debería ser 12100
        }])
        v = validar_cuadre(df)
        assert v["valido"]   # errores=0, son advertencias
        assert v["filas_con_advertencia"] == 1
        assert "991" in v["advertencias"][0]

    def test_diferencia_tolerable(self):
        # Diferencia de centavos por redondeo
        df = pd.DataFrame([{
            "Tipo":               "1",
            "Número Desde":      "1",
            "Neto Gravado Total": "10000",
            "Neto No Gravado":    "0",
            "Op. Exentas":        "0",
            "Total IVA":          "2100,00",
            "Otros Tributos":     "0",
            "Imp. Total":         "12100,00",
        }])
        v = validar_cuadre(df)
        assert v["filas_con_advertencia"] == 0


# ─── construir_hwcrarca_xlsx (integración) ────────────────────────────────────

class TestConstruirHwcrarcaXlsx:
    def _df_minimo(self) -> pd.DataFrame:
        """Crea un DataFrame de 2 filas (1 Factura A, 1 Factura B) con todas las columnas."""
        return pd.DataFrame([
            {
                "Fecha":              "01/02/2026",
                "Tipo":               "1 - Factura A",
                "Punto de Venta":     "1",
                "Número Desde":       "991",
                "Número Hasta":       "991",
                "Cód. Autorización":  "86052408575117",
                "Tipo Doc. Emisor":   "CUIT",
                "Nro. Doc. Emisor":   "30717628485",
                "Denominación Emisor": "DE CONTEINER S. A. S.",
                "Tipo Doc. Receptor": "CUIT",
                "Nro. Doc. Receptor": "30709212083",
                "Tipo Cambio":        "1",
                "Moneda":             "$",
                "Neto Grav. IVA 0%":  "0",
                "IVA 2,5%":           "0",
                "Neto Grav. IVA 2,5%": "0",
                "IVA 5%":             "0",
                "Neto Grav. IVA 5%":  "0",
                "IVA 10,5%":          "0",
                "Neto Grav. IVA 10,5%": "0",
                "IVA 21%":            "17850",
                "Neto Grav. IVA 21%": "85000",
                "IVA 27%":            "0",
                "Neto Grav. IVA 27%": "0",
                "Neto Gravado Total": "85000",
                "Neto No Gravado":    "0",
                "Op. Exentas":        "0",
                "Otros Tributos":     "0",
                "Total IVA":          "17850",
                "Imp. Total":         "102850",
            },
            {
                "Fecha":              "01/02/2026",
                "Tipo":               "11 - Factura C",
                "Punto de Venta":     "1",
                "Número Desde":       "140200",
                "Número Hasta":       "140200",
                "Cód. Autorización":  "86051918900276",
                "Tipo Doc. Emisor":   "CUIT",
                "Nro. Doc. Emisor":   "30525441969",
                "Denominación Emisor": "CAMARA ARG DE LA CONSTRUCCION",
                "Tipo Doc. Receptor": "CUIT",
                "Nro. Doc. Receptor": "30709212083",
                "Tipo Cambio":        "1",
                "Moneda":             "$",
                "Neto Grav. IVA 0%":  "0",
                "IVA 2,5%":           "0",
                "Neto Grav. IVA 2,5%": "0",
                "IVA 5%":             "0",
                "Neto Grav. IVA 5%":  "0",
                "IVA 10,5%":          "0",
                "Neto Grav. IVA 10,5%": "0",
                "IVA 21%":            "0",
                "Neto Grav. IVA 21%": "0",
                "IVA 27%":            "0",
                "Neto Grav. IVA 27%": "0",
                "Neto Gravado Total": "0",
                "Neto No Gravado":    "0",
                "Op. Exentas":        "0",
                "Otros Tributos":     "0",
                "Total IVA":          "0",
                "Imp. Total":         "62000",
            },
        ])

    def test_genera_xlsx_valido(self):
        df = self._df_minimo()
        xlsx_bytes, stats = construir_hwcrarca_xlsx(df)
        assert xlsx_bytes
        assert stats["total_filas"] == 2
        assert stats["filas_b_c"] == 1
        # Cargar y verificar estructura
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["HWComprobantes Recibidos"]

        # Fila 1: banner
        assert ws.cell(1, 1).value == BANNER_TEXT

        # Fila 2: 30 headers
        for i, header in enumerate(HWCRARCA_HEADERS, start=1):
            assert ws.cell(2, i).value == header

        # Fila 3: primera data row, Tipo extraído
        assert ws.cell(3, 2).value == "1"   # Tipo "1 - Factura A" → "1"
        assert ws.cell(3, 7).value == "80"  # Tipo Doc CUIT → 80
        assert ws.cell(3, 12).value == 1.0  # Tipo Cambio numérico

        # Cell numérica con valor (col V = Neto 21%)
        assert float(ws.cell(3, 22).value) == 85000.0

        # Fila 4: Factura C (Tipo B/C) — Neto e IVA en 0
        assert ws.cell(4, 2).value == "11"
        assert float(ws.cell(4, 22).value) == 0.0   # Neto 21% = 0
        assert float(ws.cell(4, 30).value) == 62000.0  # Imp. Total preservado

    def test_consumidor_final(self):
        df = pd.DataFrame([{
            "Fecha": "01/02/2026", "Tipo": "6", "Punto de Venta": "1",
            "Número Desde": "1", "Número Hasta": "1", "Cód. Autorización": "",
            "Tipo Doc. Emisor": "CUIT", "Nro. Doc. Emisor": "0",
            "Denominación Emisor": "",
            "Tipo Doc. Receptor": "CUIT", "Nro. Doc. Receptor": "30709212083",
            "Tipo Cambio": "1", "Moneda": "$",
            "Neto Grav. IVA 0%": "0",
            "IVA 2,5%": "0", "Neto Grav. IVA 2,5%": "0",
            "IVA 5%": "0", "Neto Grav. IVA 5%": "0",
            "IVA 10,5%": "0", "Neto Grav. IVA 10,5%": "0",
            "IVA 21%": "0", "Neto Grav. IVA 21%": "0",
            "IVA 27%": "0", "Neto Grav. IVA 27%": "0",
            "Neto Gravado Total": "0", "Neto No Gravado": "0", "Op. Exentas": "0",
            "Otros Tributos": "0", "Total IVA": "0", "Imp. Total": "5000",
        }])
        xlsx_bytes, stats = construir_hwcrarca_xlsx(df)
        assert stats["filas_consumidor_final"] == 1
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["HWComprobantes Recibidos"]
        # CUIT, Razón Social y Tipo Doc normalizados
        assert ws.cell(3, 8).value == "0"
        assert ws.cell(3, 9).value == "CONSUMIDOR FINAL"
        assert ws.cell(3, 7).value == "99"


# ─── Integración real con fixture BUTALO ──────────────────────────────────────

@pytest.mark.integration
class TestPipelineButaloFeb2026:
    """Pipeline R-01 → R-02 → R-10 con archivo real."""

    FIXTURE = Path(__file__).parent / "fixtures" / "comprobantes_butalo_feb2026.xlsx"

    def test_pipeline_completo(self):
        if not self.FIXTURE.exists():
            pytest.skip(f"Fixture no disponible: {self.FIXTURE}")

        from src.transformaciones.division_alicuotas import aplicar_division_alicuotas
        from src.transformaciones.limpieza_inicial import (
            corregir_columna_L,
            corregir_tipo_bc,
        )

        # Leer ARCA Excel (header=1)
        df = pd.read_excel(self.FIXTURE, header=1, dtype=str)

        # R-01
        df = corregir_tipo_bc(df)
        df = corregir_columna_L(df)

        # R-02
        df, _stats_r02 = aplicar_division_alicuotas(df)

        # R-10
        xlsx_bytes, stats = construir_hwcrarca_xlsx(df)

        # Asserts
        assert xlsx_bytes
        assert stats["total_filas"] == len(df)
        assert stats["validacion"]["valido"]   # sin errores

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "HWComprobantes Recibidos" in wb.sheetnames
        ws = wb["HWComprobantes Recibidos"]
        assert ws.cell(1, 1).value == BANNER_TEXT
        # Datos a partir de fila 3
        assert ws.cell(3, 1).value, "fila 3 col A debe tener fecha"
