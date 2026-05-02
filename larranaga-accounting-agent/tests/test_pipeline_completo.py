"""
E-01: Test de integración end-to-end del pipeline IVA completo.

Pipeline: ARCA crudo → R-01 (limpieza) → R-02 (división alícuotas) → R-10 (HWCRARCA)

Fixture: tests/fixtures/comprobantes_butalo_feb2026.xlsx
  - 359 filas crudas de ARCA (BUTALO SRL Feb 2026)
  - 12 facturas Tipo B/C
  - 9 filas multi-alícuota que deben expandirse

Verificaciones (todas en un único test):
  (a) Filas de entrada vs salida: 359 → 368 (9 expandidas)
  (b) Cuadre Debe = Haber a nivel agregado (con tolerancia relativa 1%)
  (c) El .xlsx generado tiene la estructura HWCRARCA correcta:
      - 30 columnas A-AD
      - Banner en F1
      - Headers en F2 con nombres exactos
      - Pestaña "HWComprobantes Recibidos"
  (d) Tipo de comprobante extraído correctamente ("1 - Factura A" → "1")
  (e) Tipo Doc mapeado ("CUIT" → "80")
  (f) Tipo Cambio numérico (1.0)
  (g) Filas Tipo B/C tienen Neto e IVA = 0, solo Imp.Total
  (h) Validación Debe=Haber pasa (cuadre_ok=True)
"""

import io
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.transformaciones.division_alicuotas import (  # noqa: E402
    aplicar_division_alicuotas,
)
from src.transformaciones.limpieza_inicial import (  # noqa: E402
    corregir_columna_L,
    corregir_tipo_bc,
    TIPOS_BC,
)
from src.transformaciones.hwcrarca_builder import (  # noqa: E402
    BANNER_TEXT,
    HWCRARCA_HEADERS,
    TIPOS_B_C as TIPOS_B_C_HWCRARCA,
    construir_hwcrarca_xlsx,
)


FIXTURE_PATH = Path(__file__).parent / "fixtures" / "comprobantes_butalo_feb2026.xlsx"


def _correr_pipeline_completo() -> tuple[bytes, dict, pd.DataFrame, pd.DataFrame]:
    """Ejecuta R-01 → R-02 → R-10 sobre el fixture y devuelve el output completo."""
    df_crudo = pd.read_excel(FIXTURE_PATH, header=1, dtype=str)
    df_post_r01 = corregir_columna_L(corregir_tipo_bc(df_crudo.copy()))
    df_post_r02, _stats_r02 = aplicar_division_alicuotas(df_post_r01)
    xlsx_bytes, stats = construir_hwcrarca_xlsx(df_post_r02)
    return xlsx_bytes, stats, df_crudo, df_post_r02


@pytest.mark.skipif(not FIXTURE_PATH.exists(), reason=f"Fixture no disponible: {FIXTURE_PATH}")
class TestE01PipelineButaloFeb2026:
    """Verificación end-to-end del pipeline ARCA → HWCRARCA con datos reales."""

    @classmethod
    def setup_class(cls):
        """Corre el pipeline una sola vez para todos los tests de la clase."""
        cls.xlsx_bytes, cls.stats, cls.df_crudo, cls.df_post_r02 = _correr_pipeline_completo()
        cls.wb = load_workbook(io.BytesIO(cls.xlsx_bytes))
        cls.ws = cls.wb["HWComprobantes Recibidos"]

    # (a) Filas de entrada vs salida
    def test_filas_entrada_359(self):
        assert len(self.df_crudo) == 359, "Fixture BUTALO debe tener 359 filas crudas"

    def test_filas_salida_368(self):
        assert len(self.df_post_r02) == 368, "R-02 debe expandir a 368 filas"

    def test_expandidas_son_9(self):
        # 368 - 359 = 9 nuevas filas por multi-alícuota
        assert len(self.df_post_r02) - len(self.df_crudo) == 9

    # (b) Cuadre Debe = Haber agregado
    def test_cuadre_ok_a_nivel_agregado(self):
        """F-10: el archivo NO debería haberse exportado si esto fallara."""
        assert self.stats["validacion"]["cuadre_ok"], (
            f"BUTALO no cuadra: Debe={self.stats['validacion']['debe_total']:.2f} "
            f"Haber={self.stats['validacion']['haber_total']:.2f}"
        )

    def test_diferencia_pct_dentro_de_tolerancia(self):
        """La diferencia debe estar dentro del 1% del haber."""
        assert self.stats["validacion"]["diferencia_pct"] < 1.0

    # (c) Estructura del .xlsx
    def test_pestaña_correcta(self):
        assert "HWComprobantes Recibidos" in self.wb.sheetnames

    def test_30_columnas(self):
        assert self.ws.max_column == 30, "HWCRARCA debe tener exactamente 30 cols A-AD"

    def test_banner_fila_1(self):
        assert self.ws.cell(1, 1).value == BANNER_TEXT

    def test_headers_fila_2_completos(self):
        for i, header_esperado in enumerate(HWCRARCA_HEADERS, start=1):
            assert self.ws.cell(2, i).value == header_esperado, (
                f"Header col {i} debe ser '{header_esperado}'"
            )

    def test_filas_de_datos_son_368(self):
        # Filas 3 a 370 = 368 filas de datos
        assert self.ws.max_row == 2 + 368

    # (d) Tipo numérico extraído
    def test_tipo_es_codigo_numerico(self):
        """Col B (Cpbte) debe ser solo el número '1', '6', '11', etc."""
        for r in range(3, self.ws.max_row + 1):
            tipo = str(self.ws.cell(r, 2).value or "")
            assert tipo.isdigit() or tipo == "", (
                f"F{r} col B Cpbte debe ser numérico, encontrado '{tipo}'"
            )

    # (e) Tipo Doc mapeado a número
    def test_tipo_doc_mapeado(self):
        """Col G (Tipo Doc) debe ser código numérico (80=CUIT, 96=DNI, 99=sin id)."""
        codigos_validos = {"80", "86", "96", "99"}
        for r in range(3, min(self.ws.max_row + 1, 30)):  # primeras 30 filas
            tipo_doc = str(self.ws.cell(r, 7).value or "")
            assert tipo_doc in codigos_validos, (
                f"F{r} col G Tipo Doc debe ser código numérico válido, encontrado '{tipo_doc}'"
            )

    # (f) Tipo Cambio numérico
    def test_tipo_cambio_es_float(self):
        """Col L (Tipo Cbio) debe ser float, no string."""
        valor = self.ws.cell(3, 12).value
        assert isinstance(valor, (int, float)), f"Tipo Cambio debe ser numérico, es {type(valor)}"
        assert valor == 1.0

    # (g) Tipo B/C: Neto e IVA en 0
    def test_tipo_b_c_neto_iva_cero(self):
        """Para todas las filas Tipo B/C, Neto Gravado e IVA deben ser 0."""
        verificadas = 0
        for r in range(3, self.ws.max_row + 1):
            tipo = str(self.ws.cell(r, 2).value or "")
            if tipo in TIPOS_B_C_HWCRARCA:
                # Col N (14) Neto Gravado, AC (29) Total IVA
                neto_gravado = self.ws.cell(r, 14).value or 0
                neto_total   = self.ws.cell(r, 25).value or 0   # Y: Imp. Neto Gravado Total
                total_iva    = self.ws.cell(r, 29).value or 0
                assert float(neto_gravado) == 0.0, f"F{r} (tipo {tipo}): Neto != 0"
                assert float(neto_total)   == 0.0, f"F{r} (tipo {tipo}): Neto Total != 0"
                assert float(total_iva)    == 0.0, f"F{r} (tipo {tipo}): IVA != 0"
                # Pero Imp.Total puede ser > 0
                imp_total = self.ws.cell(r, 30).value or 0
                if imp_total > 0:
                    verificadas += 1
        assert verificadas >= 5, "Debería haber al menos 5 filas Tipo B/C con Imp.Total > 0"

    def test_stats_filas_b_c_es_12(self):
        assert self.stats["filas_b_c"] == 12

    # (h) El builder reportó cuadre OK en stats
    def test_stats_cuadre_ok(self):
        assert self.stats["validacion"]["cuadre_ok"] is True

    # (i) Resumen para diagnóstico (no es assert, es info)
    def test_imprimir_resumen(self, capsys):
        v = self.stats["validacion"]
        print(f"\n=== RESUMEN E-01 BUTALO Feb 2026 ===")
        print(f"Filas entrada:        {len(self.df_crudo)}")
        print(f"Filas post R-02:      {len(self.df_post_r02)}")
        print(f"Filas en HWCRARCA:    {self.ws.max_row - 2}")
        print(f"Filas Tipo B/C:       {self.stats['filas_b_c']}")
        print(f"Filas consumidor fin: {self.stats['filas_consumidor_final']}")
        print(f"Suma Debe:               ${v['debe_total']:>15,.2f}")
        print(f"Suma Haber:              ${v['haber_total']:>15,.2f}")
        print(f"Diferencia:           ${v['diferencia_agregada']:>15,.2f} ({v['diferencia_pct']:.4f}%)")
        print(f"Cuadre OK:            {v['cuadre_ok']}")
        print(f"Filas con advertencia: {v['filas_con_advertencia']}")
        # Este test siempre pasa; existe solo para imprimir el resumen.
        assert True
